#!/usr/bin/env python3
"""Parse a client budget file (long format) into monthly budget lines.

A budget file has some subset of dimension columns (Brand / Region / Category)
plus one amount column. We auto-detect the columns by header name, read each
row, and normalize the amount to a monthly figure based on the declared period
(monthly | annual | total, where total is spread over `window_months`).

Returns: list of {"brand", "region", "category", "monthly"} (dims may be None).
"""
import csv
import io
import re

_HDR = {
    "brand": ("brand", "bu", "business unit", "account"),
    "region": ("region", "state", "market", "metro", "dma", "geo", "location", "territory"),
    "category": ("category", "categories", "cat", "service", "services", "product", "line", "vertical"),
    "amount": ("budget", "amount", "monthly", "annual", "spend", "target", "allocation", "$", "usd"),
}


def _slug(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def _to_num(v):
    if v is None:
        return None
    s = str(v).replace(",", "").replace("$", "").replace("%", "").strip()
    if s in ("", "-", "--"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _match_col(headers, keys):
    """Return the index of the first header matching any key (substring)."""
    slugs = [_slug(h) for h in headers]
    for i, h in enumerate(slugs):
        if any(k in h for k in keys):
            return i
    return None


def _rows_from_csv(data):
    text = data.decode("utf-8-sig", errors="replace") if isinstance(data, (bytes, bytearray)) else data
    return [r for r in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(data):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Excel (.xlsx) support needs openpyxl — please upload a CSV, or install openpyxl.")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if c is None else c for c in row])
    return out


def parse_budget_file(data, filename, period="monthly", window_months=12):
    name = (filename or "").lower()
    rows = _rows_from_xlsx(data) if name.endswith(".xlsx") else _rows_from_csv(data)
    rows = [r for r in rows if any(str(c).strip() for c in r)]   # drop blank rows
    if len(rows) < 2:
        raise ValueError("budget file has no data rows")

    # find the header row: first row where an amount column is detectable
    header_idx, cols = None, None
    for i, r in enumerate(rows[:5]):
        c = {k: _match_col(r, keys) for k, keys in _HDR.items()}
        if c["amount"] is not None and (c["brand"] is not None or c["region"] is not None or c["category"] is not None):
            header_idx, cols = i, c
            break
    if cols is None:
        raise ValueError("could not detect a Budget/Amount column and at least one of Brand/Region/Category")

    div = 12.0 if period == "annual" else (max(1, window_months) if period == "total" else 1.0)
    lines, total = [], 0.0
    for r in rows[header_idx + 1:]:
        amt = _to_num(r[cols["amount"]]) if cols["amount"] < len(r) else None
        if amt is None:
            continue
        def cell(key):
            idx = cols[key]
            if idx is None or idx >= len(r):
                return None
            v = str(r[idx]).strip()
            return v or None
        monthly = round(amt / div, 2)
        lines.append({"brand": cell("brand"), "region": cell("region"),
                      "category": cell("category"), "monthly": monthly})
        total += monthly
    if not lines:
        raise ValueError("no valid budget rows found")
    dims = [k for k in ("brand", "region", "category") if cols[k] is not None]
    return {"lines": lines, "dimensions": dims, "total_monthly": round(total, 2),
            "period": period, "count": len(lines)}
